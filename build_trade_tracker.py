"""
Build Trade Tracker
=====================
This script UPDATES your real "Trade Tracker Template 2026.xlsx" file,
adding any completed trades that aren't already recorded there. The
trades themselves now come from trading.db (see database.py) rather
than this script reading the CSV directly - database.py is what turns
your raw Fidelity export into a clean list of completed (matched
buy+sell) trades.

HOW IT WORKS: since trading.db always has your FULL trade history,
this script:
  1. Opens your existing tracker file and reads the trades already
     written there (Symbol, Date of Entry, Entry Price, # Shares,
     Date of Exit, Exit Price for each row).
  2. Compares every trade passed in against what's already in the
     tracker, skipping any that match a row that's already there.
     This is how it avoids adding duplicate rows if the database's
     trade list overlaps with trades you've already recorded.
  3. Appends only the genuinely new trades to the bottom of the
     existing list (the "Trade #" column keeps counting up from
     whatever number the tracker already ended on).
  4. Saves the result back to the SAME file, in place.

(An earlier version of this script rebuilt the whole sheet from
scratch every run, using a copy of a separate template file. That's no
longer needed now that we're appending to your real tracker directly -
but the reason we never build a sheet from a blank openpyxl.Workbook()
still applies: an early version of this script did that and Excel
flagged the output as corrupted on open. So this script still only
ever works with your real, Excel-authored file, never a blank one.)

WHAT THIS SCRIPT DOES WITH THE DATA:
  - Writes one row per new completed trade, with the same formulas
    (% change, win/loss, etc.), currency/percent/date formatting,
    centered text, and green/red color-coding the sheet already used -
    applied consistently, no matter how many rows already existed.

A NOTE ON "Portfolio Value" (column I): this is your total account
value at the time of each trade, which isn't something the CSV tells
us. Since we don't know your current account total, this script fills
in a placeholder (see LAST_KNOWN_PORTFOLIO_VALUE below) for every new
row. Change the number below, or edit column I directly in Excel
afterward, once you know the real number(s).

You can run this script directly (`python build_trade_tracker.py`) to
just refresh Excel from whatever's already in trading.db, without
importing a new CSV first. Day to day, though, run import_trades.py
instead - it imports any new CSV data AND then calls this script's
update_excel_tracker() for you.
"""

from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, PatternFill

import database

# The real tracker file this script updates - not a copy in the
# project folder, but your actual working file.
TARGET_PATH = Path(
    r"C:\Users\conno\OneDrive\Desktop\Trading\Trade Tracker Temp\Trade Tracker Template 2026.xlsx"
)
SHEET_NAME = "Trade Tracker"
TABLE_NAME = "Table1"

# See the note above - this is a placeholder since the CSV doesn't
# contain your account's total value.
LAST_KNOWN_PORTFOLIO_VALUE = 50000

# The tracker's columns, in order, starting at column A.
COLUMNS = [
    "Trade #", "Symbol", "Date of Entry", "Entry Price", "# Shares",
    "Initial Value", "Date of Exit", "Exit Price", "Portfolio Value",
    "Final Value", "Value Change", "% Change", "Equity Contribution",
    "Winning Trade", "Losing Trade", "Winning %", "Losing %",
    "Days Winning Trade", "Days Losing Trade", "Column1", "Notes",
]
FIRST_DATA_ROW = 2  # row 1 is the header

# Excel can display the same date many different ways (5/13/2026,
# 13-May-26, 2026-05-13...) depending on a cell's leftover formatting.
# We set this format explicitly on every date cell we write, so the
# whole column looks the same: month/day/year, e.g. 5/13/2026.
DATE_FORMAT = "m/d/yyyy"

# The sheet color-codes "Value Change" (K) and "% Change" / "Equity
# Contribution" (L, M) green for a win and red for a loss. These are
# the same colors the sheet already used.
GREEN_FILL = PatternFill(start_color="FFB7E1CD", end_color="FFB7E1CD", fill_type="solid")
RED_FILL = PatternFill(start_color="FFEA9999", end_color="FFEA9999", fill_type="solid")


def trade_key(symbol, entry_date, buy_price, quantity, exit_date, sell_price):
    """
    Builds a tuple that identifies one specific completed trade, used
    to tell whether a trade is already in the tracker. Prices/shares
    are rounded because a number can come back from Excel very
    slightly different (like 150.00000001) than the number we
    originally calculated from the CSV, even though it's really the
    same trade.
    """
    return (
        symbol,
        entry_date.date() if hasattr(entry_date, "date") else entry_date,
        round(buy_price, 2),
        round(quantity, 4),
        exit_date.date() if hasattr(exit_date, "date") else exit_date,
        round(sell_price, 2),
    )


def get_existing_trades(ws, table):
    """
    Reads the trades already sitting in the tracker sheet, so we know
    what NOT to re-add. Returns a set of trade_key(...) tuples for
    every already-recorded trade, plus the highest "Trade #" already
    used (0 if the tracker has no trades yet) and the row number of
    the existing "Total" row.
    """
    total_row = int(table.ref.split(":")[1][1:])

    existing_keys = set()
    max_trade_num = 0

    for row in range(FIRST_DATA_ROW, total_row):
        trade_num = ws.cell(row=row, column=1).value
        if not isinstance(trade_num, (int, float)):
            continue  # not a real trade row

        max_trade_num = max(max_trade_num, int(trade_num))
        existing_keys.add(trade_key(
            symbol=ws.cell(row=row, column=2).value,
            entry_date=ws.cell(row=row, column=3).value,
            buy_price=ws.cell(row=row, column=4).value,
            quantity=ws.cell(row=row, column=5).value,
            exit_date=ws.cell(row=row, column=7).value,
            sell_price=ws.cell(row=row, column=8).value,
        ))

    return existing_keys, max_trade_num, total_row


def get_reference_styles(ws, num_columns, reference_row=2):
    """
    Reads the number formatting (currency, percent, date, etc.) AND the
    text alignment (e.g. centered) that each column already uses on
    `reference_row`, so we can copy that same look onto every row we
    write. Row 2 is a safe row to copy from - it's part of the
    original template and we never change its formatting, only the
    numbers inside it.
    """
    formats = []
    alignments = []
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=reference_row, column=col)
        formats.append(cell.number_format)
        alignments.append(Alignment(
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical,
            wrap_text=cell.alignment.wrap_text,
        ))
    return formats, alignments


def write_trade_row(ws, row, trade_num, trade, formats, alignments):
    """Writes one completed trade into the sheet, including the same
    formulas the sheet already used for every other row, and applies
    consistent formatting (currency, percent, dates, centering) to
    every cell."""
    values = {
        1: trade_num,                                                  # Trade #
        2: trade["symbol"],                                            # Symbol
        3: trade["entry_date"],                                        # Date of Entry
        4: trade["buy_price"],                                         # Entry Price
        5: trade["quantity"],                                          # # Shares
        6: f"=E{row}*D{row}",                                          # Initial Value
        7: trade["date"],                                              # Date of Exit
        8: trade["sell_price"],                                        # Exit Price
        9: LAST_KNOWN_PORTFOLIO_VALUE,                                 # Portfolio Value
        10: f"=H{row}*E{row}",                                         # Final Value
        11: f"=J{row}-F{row}",                                         # Value Change
        12: f"=ROUND(((H{row}/D{row})-1)*100,4)/100",                  # % Change
        13: f"=K{row}/Table1[[#This Row],[Portfolio Value]]",          # Equity Contribution
        14: f"=IF(L{row}>0,1,0)",                                      # Winning Trade
        15: f"=IF(N{row}>0,0,1)",                                      # Losing Trade
        16: f"=IF(N{row}=1,L{row},0)*100",                             # Winning %
        17: f"=IF(O{row}=0,0,L{row})*100",                             # Losing %
        18: f"=IF(N{row}=1,G{row}-C{row},0)",                          # Days Winning Trade
        19: f"=IF(N{row}=1,0,G{row}-C{row})",                          # Days Losing Trade
        20: f"=Table1[[#This Row],[Date of Exit]]-Table1[[#This Row],[Date of Entry]]",  # Column1
        # 21 (Notes) is left blank for you to fill in
    }

    for col, value in values.items():
        cell = ws.cell(row=row, column=col, value=value)
        cell.number_format = formats[col - 1]
        cell.alignment = alignments[col - 1]

    # Dates get their own explicit format so they always read
    # month/day/year, regardless of what the reference row used.
    ws.cell(row=row, column=3).number_format = DATE_FORMAT
    ws.cell(row=row, column=7).number_format = DATE_FORMAT


def write_totals_row(ws, row, formats, alignments):
    """Writes the 'Total' row underneath the trade data, using the same
    SUBTOTAL formulas the sheet already had."""
    totals = {
        1: "Total",
        11: "=SUBTOTAL(109,Table1[Value Change])",
        13: "=SUBTOTAL(109,Table1[Equity Contribution])",
        14: "=SUBTOTAL(109,Table1[Winning Trade])",
        15: "=SUBTOTAL(109,Table1[Losing Trade])",
        16: "=SUBTOTAL(109,Table1[Winning %])",
        17: "=SUBTOTAL(109,Table1[Losing %])",
        18: "=SUBTOTAL(109,Table1[Days Winning Trade])",
        19: "=SUBTOTAL(109,Table1[Days Losing Trade])",
    }
    for col, value in totals.items():
        cell = ws.cell(row=row, column=col, value=value)
        cell.number_format = formats[col - 1]
        cell.alignment = alignments[col - 1]


def refresh_conditional_formatting(ws, last_trade_row):
    """
    The sheet color-codes some columns (green for a gain, red for a
    loss), but those color rules were only ever set up to cover a
    fixed, hand-picked set of rows in the original template - they
    don't automatically grow when more trades are added.

    This removes those old, out-of-date color rules for columns K
    (Value Change) and L:M (% Change, Equity Contribution), and adds
    fresh ones that cover every row we just wrote.
    """
    stale_ranges = [
        rng for rng in ws.conditional_formatting._cf_rules
        if str(rng.sqref).startswith("K2") or str(rng.sqref).startswith("L2")
    ]
    for rng in stale_ranges:
        del ws.conditional_formatting._cf_rules[rng]

    for column_range in (f"K2:K{last_trade_row}", f"L2:M{last_trade_row}"):
        ws.conditional_formatting.add(
            column_range, CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL))
        ws.conditional_formatting.add(
            column_range, CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))


def update_excel_tracker(all_trades):
    """
    Brings Trade Tracker Template 2026.xlsx up to date with the given
    list of completed trades (same dict shape database.get_trades()
    returns), appending only the ones that aren't already in the sheet.
    """
    # The sheet's own formulas (% Change = exit/entry - 1, Value Change
    # = final - initial, etc.) are all built on the buy-low-sell-high
    # assumption of a LONG trade - a short trade's numbers would come
    # out with the wrong sign no matter which way its prices were
    # written into the columns. So short trades are left out of Excel
    # entirely (they're still fully tracked in the database and on
    # every dashboard page).
    short_count = sum(1 for t in all_trades if t.get("direction") == "SHORT")
    if short_count:
        print(f"Note: {short_count} short trade(s) not written to Excel - the "
              "template's formulas assume long trades.")
    all_trades = [t for t in all_trades if t.get("direction") != "SHORT"]

    wb = openpyxl.load_workbook(TARGET_PATH)
    ws = wb[SHEET_NAME]
    table = ws.tables[TABLE_NAME]

    # Read what's already in the tracker, so we know which of the
    # given trades are actually new, and what "Trade #" to continue
    # counting from.
    existing_keys, max_trade_num, old_total_row = get_existing_trades(ws, table)

    new_trades = [
        t for t in all_trades
        if trade_key(t["symbol"], t["entry_date"], t["buy_price"], t["quantity"],
                     t["date"], t["sell_price"]) not in existing_keys
    ]

    if not new_trades:
        print("No new trades to add - the tracker is already up to date.")
        return

    print(f"Adding {len(new_trades)} new trade(s) to the tracker.")

    # Capture the formatting (currency/percent/date styles, and
    # centering) from row 2 before we touch anything, so every row we
    # write can match it.
    formats, alignments = get_reference_styles(ws, len(COLUMNS))

    # The old "Total" row is about to become the first new trade row,
    # so new rows start exactly where it was.
    for i, trade in enumerate(new_trades):
        row = old_total_row + i
        write_trade_row(ws, row, max_trade_num + 1 + i, trade, formats, alignments)

    totals_row = old_total_row + len(new_trades)
    write_totals_row(ws, totals_row, formats, alignments)

    # Tell Excel the table now covers this new range, so formulas,
    # filtering, and formatting keep working correctly.
    table.ref = f"A1:U{totals_row}"
    table.autoFilter.ref = f"A1:U{totals_row - 1}"

    # Make the green/red color-coding cover every row, old and new.
    refresh_conditional_formatting(ws, totals_row - 1)

    wb.save(TARGET_PATH)
    print(f"Done! Updated: {TARGET_PATH.name}")


def main():
    """Refreshes Excel from whatever's already stored in trading.db,
    without importing any new CSV data first."""
    conn = database.get_connection()
    all_trades = database.get_trades(conn)
    print(f"Found {len(all_trades)} completed stock trades in trading.db.")
    update_excel_tracker(all_trades)


if __name__ == "__main__":
    main()
